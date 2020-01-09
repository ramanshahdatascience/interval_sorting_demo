#! /usr/bin/env python


import csv
from openpyxl import load_workbook

workbook = load_workbook(filename='./downloads/townest.xlsx')
with open('./transformed/population.csv', 'w', newline='') as csvfile:
    fieldnames = ['municipality', 'population']
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()

    for row in workbook.active.iter_rows(min_row=9, max_row=52, min_col=1, max_col=2):
        if 'County' not in row[0].value:
            writer.writerow({'municipality': row[0].value, 'population': row[1].value})
